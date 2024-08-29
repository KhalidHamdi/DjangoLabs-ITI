from django.shortcuts import render
from django.http import HttpResponseRedirect
from openpyxl import load_workbook

EXCEL_FILE = 'tracking_system/data/orders.xlsx'

def home(request):
    return render(request, 'home.html')

def read_excel():
    wb = load_workbook(EXCEL_FILE)
    ws = wb.active
    orders = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        order = {
            'orderID': row[0],
            'orderName': row[1],
            'username': row[2],
            'status': row[3],
        }
        orders.append(order)
    return orders

def tracking(request):
    if request.method == 'POST':
        data = request.POST
        wb = load_workbook(EXCEL_FILE)
        ws = wb.active
        default_status = "requested"
        ws.append([data['orderID'], data['orderName'], data['username'], default_status])
        wb.save(EXCEL_FILE)
        return HttpResponseRedirect('/tracking/')  

    orders = read_excel()
    return render(request, 'tracking.html', {'orders': orders})
